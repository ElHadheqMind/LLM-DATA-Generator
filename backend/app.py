#!/usr/bin/env python3
"""
Flask backend for LLM Data Generator
Handles PDF upload and processing
"""

import os
import tempfile
import base64
import logging
import threading
import time
import json
from datetime import datetime
from pathlib import Path

from flask import Flask, request, jsonify
from flask_cors import CORS
from werkzeug.utils import secure_filename
from werkzeug.exceptions import RequestEntityTooLarge
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill
import aiohttp

# Import our document extractors and AI services
from document_extractor import DocumentExtractor
from dotenv import load_dotenv
# AI service manager and configuration loader
from ai_service_manager import AIServiceManager
from ai_config_loader import load_all_ai_configs
import asyncio

# Load environment variables
load_dotenv()
# Force reload to pick up Azure OpenAI credentials

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def extract_json_from_text(text: str) -> str:
    """
    Extract JSON object from text by finding the first { and matching }.

    Args:
        text (str): Text that may contain JSON

    Returns:
        Optional[str]: Extracted JSON string or None if not found
    """
    # Find the first opening brace
    start_idx = text.find('{')
    if start_idx == -1:
        return None

    # Count braces to find the matching closing brace
    brace_count = 0
    in_string = False
    escape_next = False

    for i, char in enumerate(text[start_idx:], start_idx):
        if escape_next:
            escape_next = False
            continue

        if char == '\\':
            escape_next = True
            continue

        if char == '"' and not escape_next:
            in_string = not in_string
            continue

        if not in_string:
            if char == '{':
                brace_count += 1
            elif char == '}':
                brace_count -= 1
                if brace_count == 0:
                    # Found the matching closing brace
                    return text[start_idx:i+1]

    return None

app = Flask(__name__)
CORS(app)  # Enable CORS for React frontend

# Initialize AI services
ai_service_manager = None
ai_initialization_complete = False
ai_initialization_lock = threading.Lock()

async def initialize_ai_services():
    """Initialize all AI services asynchronously."""
    global ai_service_manager, ai_initialization_complete

    try:
        logger.info("Initializing AI services...")

        # Always create AI service manager (no longer dependent on QUESTION_GENERATION_ENABLED)
        ai_service_manager = AIServiceManager()
        configs = load_all_ai_configs()

        if configs:
            logger.info(f"Initializing {len(configs)} AI providers...")
            results = await ai_service_manager.initialize_providers(configs)

            successful_providers = [provider.value for provider, success in results.items() if success]
            failed_providers = [provider.value for provider, success in results.items() if not success]

            if successful_providers:
                logger.info(f"✅ Successfully initialized providers: {', '.join(successful_providers)}")
            if failed_providers:
                logger.warning(f"❌ Failed to initialize providers: {', '.join(failed_providers)}")

            # Keep the service manager even if no providers initialized
            # This allows on-demand initialization when credentials are added
        else:
            logger.info("No AI provider configurations found yet - waiting for credentials")
            # Keep the service manager for on-demand initialization
    except Exception as e:
        logger.error(f"Error initializing AI services: {str(e)}")
        # Still create the service manager for on-demand initialization
        if ai_service_manager is None:
            ai_service_manager = AIServiceManager()
    finally:
        # Mark initialization as complete regardless of success/failure
        with ai_initialization_lock:
            ai_initialization_complete = True
        logger.info("AI services initialization complete")

# Initialize AI services in a separate thread to avoid blocking Flask startup
def init_ai_services_sync():
    """Synchronous wrapper for AI service initialization."""
    global ai_service_manager, ai_initialization_complete

    # Clean up existing services before reinitializing (important for hot reload)
    if ai_service_manager:
        try:
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            loop.run_until_complete(ai_service_manager.cleanup())
            loop.close()
        except Exception as e:
            logger.warning(f"Error cleaning up existing AI services: {e}")

    try:
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        loop.run_until_complete(initialize_ai_services())
        loop.close()
    except Exception as e:
        logger.error(f"Failed to initialize AI services: {e}")
        # Mark as complete even on error
        with ai_initialization_lock:
            ai_initialization_complete = True

# Start AI service initialization in background
ai_init_thread = threading.Thread(target=init_ai_services_sync, daemon=True)
ai_init_thread.start()

# Configuration
UPLOAD_FOLDER = tempfile.gettempdir()
# Get supported extensions from DocumentExtractor
ALLOWED_EXTENSIONS = DocumentExtractor.get_supported_extensions()
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = MAX_FILE_SIZE

def allowed_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def sanitize_excel_string(text):
    """
    Sanitize string for Excel compatibility by removing or replacing problematic characters.
    Excel has limitations on certain control characters and very long strings.

    Args:
        text (str): Text to sanitize

    Returns:
        str: Sanitized text safe for Excel
    """
    if not text:
        return text

    # Convert to string if not already
    text = str(text)

    # Remove control characters that Excel doesn't like (except newline, tab, carriage return)
    # Keep printable characters and common whitespace
    sanitized = ''.join(char for char in text if ord(char) >= 32 or char in '\n\r\t')

    # Excel has a cell limit of 32,767 characters
    if len(sanitized) > 32767:
        sanitized = sanitized[:32764] + '...'

    return sanitized


def create_qa_excel_file(qa_data, output_path):
    """
    Create an Excel file with Question-Answer pairs format.
    Supports both single-document and multi-document formats.

    Args:
        qa_data (List[Dict]): List of data items with questions and hierarchical content
        output_path (str): Path where the Excel file should be saved
    """
    # Check if this is multi-document data (has source_document field)
    has_multi_documents = any(item.get('source_document') for item in qa_data)

    # Prepare data for Excel
    excel_data = []

    for item in qa_data:
        question = sanitize_excel_string(item.get('question', ''))

        # Use generated answer if available, otherwise use content
        answer = sanitize_excel_string(item.get('answer', item.get('content', '')))

        # Create full hierarchical content by concatenating all information
        content_parts = []

        # Add section hierarchy
        if item.get('section'):
            content_parts.append(f"Section: {item['section']}")
        if item.get('subsection'):
            content_parts.append(f"Subsection: {item['subsection']}")
        if item.get('subsubsection'):
            content_parts.append(f"Sub-subsection: {item['subsubsection']}")
        if item.get('subsubsubsection'):
            content_parts.append(f"Sub-sub-subsection: {item['subsubsubsection']}")

        # Add the actual content
        if item.get('content'):
            content_parts.append(f"Content: {item['content']}")

        # Join all parts with separators for full context
        full_content = sanitize_excel_string(' | '.join(content_parts))

        # Create row data based on whether this is multi-document or single-document
        if has_multi_documents:
            # 4-column format: Document Name, Question, Answer/Context, Full Content
            excel_data.append({
                'Document Name': sanitize_excel_string(item.get('source_document', 'Unknown Document')),
                'Question': question,
                'Answer': answer,
                'Context': full_content
            })
        else:
            # 3-column format: Question, Answer, Content (backward compatibility)
            excel_data.append({
                'Question': question,
                'Answer': answer,
                'Content': full_content
            })

    # Create DataFrame
    df = pd.DataFrame(excel_data)

    # Save to Excel with formatting
    try:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='QA_Dataset', index=False)

            # Get the worksheet for formatting
            worksheet = writer.sheets['QA_Dataset']

            # Style the header row
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')

            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Auto-adjust column widths with specific handling for multi-document format
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                column_header = worksheet[f'{column_letter}1'].value

                for cell in column:
                    try:
                        # Handle multi-line content
                        if cell.value:
                            lines = str(cell.value).split('\n')
                            max_line_length = max(len(line) for line in lines) if lines else 0
                            if max_line_length > max_length:
                                max_length = max_line_length
                    except Exception as e:
                        logger.warning(f"Error calculating column width for cell: {e}")
                        pass

                # Set column width with specific limits based on column type
                if column_header == 'Document Name':
                    # Document name column - moderate width
                    adjusted_width = min(max(max_length + 2, 20), 40)
                elif column_header in ['Question', 'Answer']:
                    # Question and Answer columns - wider
                    adjusted_width = min(max(max_length + 2, 30), 60)
                elif column_header in ['Content', 'Context']:
                    # Content/Context column - widest
                    adjusted_width = min(max(max_length + 2, 40), 80)
                else:
                    # Default width
                    adjusted_width = min(max(max_length + 2, 15), 80)

                worksheet.column_dimensions[column_letter].width = adjusted_width

            # Set row height for better readability of multi-line content
            for row in worksheet.iter_rows(min_row=2):  # Skip header row
                for cell in row:
                    try:
                        if cell.value and '\n' in str(cell.value):
                            # Calculate approximate height based on number of lines
                            line_count = str(cell.value).count('\n') + 1
                            worksheet.row_dimensions[cell.row].height = max(15 * line_count, 30)

                            # Enable text wrapping
                            cell.alignment = Alignment(wrap_text=True, vertical='top')
                    except Exception as e:
                        logger.warning(f"Error setting row height for cell: {e}")
                        pass

        logger.info(f"Q&A Excel file saved: {output_path}")
        logger.info(f"Total Q&A pairs: {len(excel_data)}")
    except Exception as e:
        logger.error(f"Error writing Excel file: {e}")
        logger.error(f"Excel data sample: {excel_data[0] if excel_data else 'No data'}")
        raise



@app.errorhandler(RequestEntityTooLarge)
def handle_file_too_large(_):
    """Handle file size too large error"""
    return jsonify({
        'error': 'File too large. Maximum file size is 10MB.',
        'success': False
    }), 413

@app.errorhandler(Exception)
def handle_general_error(e):
    """Handle general errors"""
    logger.error(f'Unhandled error: {str(e)}')
    return jsonify({
        'error': f'Internal server error: {str(e)}',
        'success': False
    }), 500

@app.route('/api/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    return jsonify({
        'status': 'healthy',
        'message': 'LLM Data Generator Backend is running',
        'version': '1.0.0'
    })

@app.route('/api/process-pdf', methods=['POST'])
def process_pdf():
    """Process uploaded document and extract hierarchical data"""
    try:
        # Check if file is present
        if 'pdf' not in request.files:
            return jsonify({
                'error': 'No document file provided',
                'success': False
            }), 400

        file = request.files['pdf']

        # Check if file is selected
        if file.filename == '':
            return jsonify({
                'error': 'No file selected',
                'success': False
            }), 400

        # Validate file type
        if not allowed_file(file.filename):
            supported_formats = ', '.join(sorted(ALLOWED_EXTENSIONS))
            return jsonify({
                'error': f'Invalid file type. Supported formats: {supported_formats}',
                'success': False
            }), 400
        
        # Save uploaded file temporarily
        filename = secure_filename(file.filename)
        temp_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(temp_path)

        file_ext = Path(filename).suffix.lower().lstrip('.')
        logger.info(f'Processing {file_ext.upper()} document: {filename}')

        try:
            # Get clean option from form data (default to False)
            clean_output = request.form.get('clean_output', 'false').lower() == 'true'

            # Get extraction method from form data (default to 'auto')
            extraction_method = request.form.get('extraction_method', 'auto')
            logger.info(f'Extraction method: {extraction_method}')

            # Use DocumentExtractor for all file types
            extractor = DocumentExtractor()
            pages_text = extractor.extract_text(temp_path)

            # Extract document name from filename (without extension)
            document_name = Path(filename).stem

            # Pass file_path for PDF TOC extraction and extraction method
            hierarchy_data = extractor.process_text_to_hierarchy(
                pages_text,
                clean_output,
                document_name,
                file_path=temp_path,
                extraction_method=extraction_method
            )

            # Clean up temporary file
            os.remove(temp_path)

            logger.info(f'Successfully processed {filename}: {len(hierarchy_data)} entries extracted (clean_output={clean_output})')

            # Return extracted data
            return jsonify({
                'success': True,
                'data': hierarchy_data,
                'filename': filename,
                'total_rows': len(hierarchy_data),
                'message': f'Successfully extracted {len(hierarchy_data)} entries from {filename}',
                'clean_applied': clean_output
            })
            
        except Exception as e:
            # Clean up temporary file on error
            if os.path.exists(temp_path):
                os.remove(temp_path)
            raise e
            
    except Exception as e:
        logger.error(f'Error processing PDF: {str(e)}')
        return jsonify({
            'error': f'Failed to process PDF: {str(e)}',
            'success': False
        }), 500

@app.route('/api/download-excel', methods=['POST'])
def download_excel():
    """Generate and return Excel file from extracted data"""
    try:
        data = request.json.get('data', [])
        filename = request.json.get('filename', 'extracted_data')

        if not data:
            return jsonify({'error': 'No data provided'}), 400

        # Create temporary Excel file
        extractor = DocumentExtractor()
        temp_excel_path = os.path.join(tempfile.gettempdir(), f"{filename}_hierarchy.xlsx")
        extractor.save_to_excel(data, temp_excel_path)

        # Read file and return as base64
        with open(temp_excel_path, 'rb') as f:
            excel_data = f.read()

        # Clean up temporary file
        os.remove(temp_excel_path)

        # Return base64 encoded file
        excel_base64 = base64.b64encode(excel_data).decode('utf-8')

        return jsonify({
            'success': True,
            'file_data': excel_base64,
            'filename': f"{filename}_hierarchy.xlsx",
            'message': 'Excel file generated successfully'
        })

    except Exception as e:
        logger.error(f'Error generating Excel file: {str(e)}')
        return jsonify({
            'error': f'Failed to generate Excel file: {str(e)}',
            'success': False
        }), 500



@app.route('/api/download-qa-excel', methods=['POST'])
def download_qa_excel():
    """Generate and return Excel file with Question-Answer pairs format"""
    try:
        data = request.json.get('data', [])
        filename = request.json.get('filename', 'qa_dataset')

        if not data:
            return jsonify({'error': 'No data provided'}), 400

        # Filter data to only include items with generated questions
        qa_data = [item for item in data if item.get('question') and item.get('question_generated')]

        if not qa_data:
            return jsonify({'error': 'No question-answer pairs found in the provided data'}), 400

        # Check if this is multi-document data
        has_multi_documents = any(item.get('source_document') for item in qa_data)
        unique_documents = set(item.get('source_document', 'Unknown') for item in qa_data if item.get('source_document'))

        logger.info(f'Generating Q&A Excel with {len(qa_data)} pairs from {len(unique_documents) if has_multi_documents else 1} document(s)')

        # Create temporary Excel file
        temp_excel_path = os.path.join(tempfile.gettempdir(), f"{filename}_qa_dataset.xlsx")
        create_qa_excel_file(qa_data, temp_excel_path)

        # Read file and return as base64
        with open(temp_excel_path, 'rb') as f:
            excel_data = f.read()

        # Clean up temporary file
        os.remove(temp_excel_path)

        # Return base64 encoded file
        excel_base64 = base64.b64encode(excel_data).decode('utf-8')

        # Create response message
        if has_multi_documents:
            message = f'Multi-document Q&A Excel file generated with {len(qa_data)} pairs from {len(unique_documents)} documents'
        else:
            message = f'Q&A Excel file generated successfully with {len(qa_data)} pairs'

        return jsonify({
            'success': True,
            'file_data': excel_base64,
            'filename': f"{filename}_qa_dataset.xlsx",
            'message': message,
            'total_qa_pairs': len(qa_data),
            'document_count': len(unique_documents) if has_multi_documents else 1,
            'is_multi_document': has_multi_documents,
            'documents': list(unique_documents) if has_multi_documents else []
        })

    except Exception as e:
        logger.error(f'Error generating Q&A Excel file: {str(e)}')
        return jsonify({
            'error': f'Failed to generate Q&A Excel file: {str(e)}',
            'success': False
        }), 500


@app.route('/api/download-qa-csv', methods=['POST'])
def download_qa_csv():
    """Generate and return CSV file with Question-Answer pairs format"""
    try:
        data = request.json.get('data', [])
        filename = request.json.get('filename', 'qa_dataset')

        if not data:
            return jsonify({'error': 'No data provided'}), 400

        # Filter data to only include items with generated questions
        qa_data = [item for item in data if item.get('question') and item.get('question_generated')]

        if not qa_data:
            return jsonify({'error': 'No question-answer pairs found in the provided data'}), 400

        # Check if this is multi-document data
        has_multi_documents = any(item.get('source_document') for item in qa_data)
        unique_documents = set(item.get('source_document', 'Unknown') for item in qa_data if item.get('source_document'))

        logger.info(f'Generating Q&A CSV with {len(qa_data)} pairs from {len(unique_documents) if has_multi_documents else 1} document(s)')

        # Prepare data for CSV
        csv_data = []
        for item in qa_data:
            question = sanitize_excel_string(item.get('question', ''))
            answer = sanitize_excel_string(item.get('answer', item.get('content', '')))

            # Create full hierarchical content
            content_parts = []
            if item.get('section'):
                content_parts.append(f"Section: {item['section']}")
            if item.get('subsection'):
                content_parts.append(f"Subsection: {item['subsection']}")
            if item.get('subsubsection'):
                content_parts.append(f"Sub-subsection: {item['subsubsection']}")
            if item.get('subsubsubsection'):
                content_parts.append(f"Sub-sub-subsection: {item['subsubsubsection']}")
            if item.get('content'):
                content_parts.append(f"Content: {item['content']}")

            full_content = sanitize_excel_string(' | '.join(content_parts))

            if has_multi_documents:
                csv_data.append({
                    'Document Name': sanitize_excel_string(item.get('source_document', 'Unknown Document')),
                    'Question': question,
                    'Answer': answer,
                    'Context': full_content
                })
            else:
                csv_data.append({
                    'Question': question,
                    'Answer': answer,
                    'Content': full_content
                })

        # Create DataFrame and convert to CSV
        df = pd.DataFrame(csv_data)
        csv_string = df.to_csv(index=False)

        # Encode as base64
        csv_base64 = base64.b64encode(csv_string.encode('utf-8')).decode('utf-8')

        # Create response message
        if has_multi_documents:
            message = f'Multi-document Q&A CSV file generated with {len(qa_data)} pairs from {len(unique_documents)} documents'
        else:
            message = f'Q&A CSV file generated successfully with {len(qa_data)} pairs'

        return jsonify({
            'success': True,
            'file_data': csv_base64,
            'filename': f"{filename}_qa_dataset.csv",
            'message': message,
            'total_qa_pairs': len(qa_data),
            'document_count': len(unique_documents) if has_multi_documents else 1,
            'is_multi_document': has_multi_documents,
            'documents': list(unique_documents) if has_multi_documents else []
        })

    except Exception as e:
        logger.error(f'Error generating Q&A CSV file: {str(e)}')
        return jsonify({
            'error': f'Failed to generate Q&A CSV file: {str(e)}',
            'success': False
        }), 500


@app.route('/api/download-qa-json', methods=['POST'])
def download_qa_json():
    """Generate and return JSON file with Question-Answer pairs format"""
    try:
        data = request.json.get('data', [])
        filename = request.json.get('filename', 'qa_dataset')

        if not data:
            return jsonify({'error': 'No data provided'}), 400

        # Filter data to only include items with generated questions
        qa_data = [item for item in data if item.get('question') and item.get('question_generated')]

        if not qa_data:
            return jsonify({'error': 'No question-answer pairs found in the provided data'}), 400

        # Check if this is multi-document data
        has_multi_documents = any(item.get('source_document') for item in qa_data)
        unique_documents = set(item.get('source_document', 'Unknown') for item in qa_data if item.get('source_document'))

        logger.info(f'Generating Q&A JSON with {len(qa_data)} pairs from {len(unique_documents) if has_multi_documents else 1} document(s)')

        # Prepare data for JSON
        json_data = []
        for item in qa_data:
            qa_item = {
                'question': item.get('question', ''),
                'answer': item.get('answer', item.get('content', '')),
                'metadata': {
                    'section': item.get('section', ''),
                    'subsection': item.get('subsection', ''),
                    'subsubsection': item.get('subsubsection', ''),
                    'subsubsubsection': item.get('subsubsubsection', ''),
                    'content': item.get('content', '')
                }
            }

            if has_multi_documents:
                qa_item['source_document'] = item.get('source_document', 'Unknown Document')

            json_data.append(qa_item)

        # Create JSON string with pretty formatting
        json_string = json.dumps({
            'qa_pairs': json_data,
            'metadata': {
                'total_pairs': len(qa_data),
                'is_multi_document': has_multi_documents,
                'document_count': len(unique_documents) if has_multi_documents else 1,
                'documents': list(unique_documents) if has_multi_documents else [],
                'generated_at': datetime.now().isoformat()
            }
        }, indent=2, ensure_ascii=False)

        # Create response message
        if has_multi_documents:
            message = f'Multi-document Q&A JSON file generated with {len(qa_data)} pairs from {len(unique_documents)} documents'
        else:
            message = f'Q&A JSON file generated successfully with {len(qa_data)} pairs'

        return jsonify({
            'success': True,
            'file_data': json_string,  # Return as string, not base64
            'filename': f"{filename}_qa_dataset.json",
            'message': message,
            'total_qa_pairs': len(qa_data),
            'document_count': len(unique_documents) if has_multi_documents else 1,
            'is_multi_document': has_multi_documents,
            'documents': list(unique_documents) if has_multi_documents else []
        })

    except Exception as e:
        logger.error(f'Error generating Q&A JSON file: {str(e)}')
        return jsonify({
            'error': f'Failed to generate Q&A JSON file: {str(e)}',
            'success': False
        }), 500


# System prompt configuration - using environment variable only

# System prompt API endpoint
@app.route('/api/system-prompt', methods=['GET'])
def get_system_prompt_endpoint():
    """Get current system prompt from environment variable"""
    try:
        from ai_service_base import get_current_system_prompt
        system_prompt = get_current_system_prompt()

        # Return 200 even if empty - frontend will handle gracefully
        return jsonify({
            'success': True,
            'system_prompt': system_prompt if system_prompt else '',
            'note': 'No system prompt configured' if not system_prompt else None
        })
    except Exception as e:
        logger.error(f'Error getting system prompt: {str(e)}')
        return jsonify({
            'success': False,
            'error': f'Failed to get system prompt: {str(e)}',
            'system_prompt': ''  # No hardcoded fallback
        }), 500

@app.route('/api/generate-system-prompt', methods=['POST'])
def generate_system_prompt_endpoint():
    """Generate a custom system prompt based on user's use case description using selected AI provider and model"""
    try:
        # Sync credentials from request headers to credential manager (in-memory only)
        from request_credentials import sync_request_credentials_to_manager
        sync_request_credentials_to_manager(request)

        data = request.json
        use_case_description = data.get('use_case_description', '')
        provider_id = data.get('provider_id', '')
        model_name = data.get('model_name', '')
        generation_mode = data.get('generation_mode', 'qa_pair')  # NEW: Get generation mode

        logger.info(f"=== System Prompt Generation Request ===")
        logger.info(f"Provider: {provider_id}, Model: {model_name}")
        logger.info(f"Generation Mode: {generation_mode}")
        logger.info(f"Use case: {use_case_description[:100]}...")

        if not use_case_description or not use_case_description.strip():
            return jsonify({
                'success': False,
                'error': 'Use case description is required'
            }), 400

        # Use the global AI service manager
        global ai_service_manager
        from ai_service_base import AIProviderType
        import asyncio

        # Check if AI service manager is initialized
        if not ai_service_manager:
            logger.error("AI service manager not initialized")
            return jsonify({
                'success': False,
                'error': 'AI services not initialized. Please check backend configuration.',
                'system_prompt': ''
            }), 500

        # Meta-system prompt: Instructions for the AI on how to create SHORT, CONCISE system prompts
        meta_system_prompt = """You are an expert at creating SHORT, CONCISE system prompts for AI models. Your task is to generate brief, focused system prompts that follow a specific template structure.

CRITICAL REQUIREMENTS:
1. Keep the prompt SHORT (150-200 characters maximum)
2. Follow the EXACT structure of the provided template
3. Only add the user's domain/context to the base template
4. Do NOT add extra instructions or make it longer
5. Output ONLY the system prompt text, nothing else

Generate a SHORT system prompt by taking the base template and incorporating the user's specific domain/use case."""

        # Meta-user prompt: Template for requesting a custom system prompt (MODE-SPECIFIC)
        if generation_mode == 'question_only':
            # Generate a SHORT Question-Only mode system prompt
            user_prompt = f"""Generate a SHORT system prompt for: {use_case_description}

BASE TEMPLATE (DO NOT CHANGE THE STRUCTURE):
"You are an expert in generating training data for LLM fine-tuning. Generate a clear, specific question based on the provided document content. Questions should cover ALL content comprehensively. Return ONLY the question text with no formatting or explanations."

YOUR TASK:
Take the base template above and add the user's domain/context to make it more specific. For example, if the use case is about "medical research papers", modify it to:
"You are an expert in generating training data for LLM fine-tuning, specializing in medical research papers. Generate a clear, specific question based on the provided document content. Questions should cover ALL content comprehensively. Return ONLY the question text with no formatting or explanations."

RULES:
- Keep it SHORT (similar length to the base template)
- Only add domain specificity where it makes sense (usually after "fine-tuning")
- Do NOT add extra instructions
- Do NOT make it longer than necessary
- Output ONLY the modified prompt text

Generate the SHORT system prompt now:"""

        else:
            # Generate a SHORT Question+Answer mode system prompt
            user_prompt = f"""Generate a SHORT system prompt for: {use_case_description}

BASE TEMPLATE (DO NOT CHANGE THE STRUCTURE):
"You are an expert in generating training data for LLM fine-tuning. Generate a question-answer pair based on the provided document content. Cover ALL content comprehensively. CRITICAL: Respond with ONLY valid JSON - no markdown, no code blocks. Format: {{"question": "...", "answer": "..."}}. The answer should be complete and detailed. Return ONLY this JSON object."

YOUR TASK:
Take the base template above and add the user's domain/context to make it more specific. For example, if the use case is about "medical research papers", modify it to:
"You are an expert in generating training data for LLM fine-tuning, specializing in medical research papers. Generate a question-answer pair based on the provided document content. Cover ALL content comprehensively. CRITICAL: Respond with ONLY valid JSON - no markdown, no code blocks. Format: {{"question": "...", "answer": "..."}}. The answer should be complete and detailed. Return ONLY this JSON object."

RULES:
- Keep it SHORT (similar length to the base template)
- Only add domain specificity where it makes sense (usually after "fine-tuning")
- Do NOT add extra instructions
- Do NOT make it longer than necessary
- Output ONLY the modified prompt text

Generate the SHORT system prompt now:"""

        async def generate():
            try:
                # Ensure the requested provider is initialized with request credentials
                if provider_id:
                    provider_type = AIProviderType(provider_id)

                    # Get credentials from request headers
                    from request_credentials import extract_credentials_from_headers
                    from ai_service_base import AIServiceConfig

                    creds = extract_credentials_from_headers(request, provider_id)
                    if creds:
                        # Create config from request credentials
                        # Only pass parameters that AIServiceConfig accepts
                        config = AIServiceConfig(
                            provider_type=provider_type,
                            api_key=creds.api_key,
                            endpoint=creds.endpoint,
                            model_name=model_name or creds.model_name,
                            deployment_name=creds.deployment_name,
                            api_version=creds.api_version
                        )

                        # Ensure provider is initialized with these credentials
                        logger.info(f"Ensuring {provider_id} is initialized with request credentials")
                        await ai_service_manager.ensure_provider_initialized(provider_type, config)

                # Get available providers (returns a list)
                providers = ai_service_manager.get_available_providers()
                if not providers:
                    logger.error("No AI providers available")
                    return None, "No AI providers available"

                # Determine provider
                if provider_id:
                    provider_type = AIProviderType(provider_id)
                    if provider_type not in providers:
                        logger.warning(f"Requested provider {provider_id} not available, using first available")
                        provider_type = providers[0]
                else:
                    provider_type = providers[0]

                logger.info(f"Using provider: {provider_type.value}")

                # Get service
                service = ai_service_manager._services.get(provider_type)
                if not service:
                    return None, f"Service not found for {provider_type.value}"

                # Set model temporarily
                original_model = service.config.model_name
                if model_name:
                    service.config.model_name = model_name

                try:
                    # Check if service has generate_text_direct method (for bypassing system prompt requirement)
                    if hasattr(service, 'generate_text_direct'):
                        # Use generate_text_direct to bypass system prompt requirement
                        # This is a meta-task: using AI to generate a system prompt
                        response = await service.generate_text_direct(
                            prompt=user_prompt,
                            system_instruction=meta_system_prompt
                        )
                    else:
                        # Fallback: Use generate_question_answer_pair with system prompt
                        # (This will work for services that don't validate system prompt)
                        response = await service.generate_question_answer_pair(
                            content=user_prompt,
                            context=None,
                            system_prompt=meta_system_prompt,
                            generation_mode='question_only'
                        )

                    if response and response.success:
                        # The response content should be the generated system prompt
                        return response.content, None
                    else:
                        return None, response.error if response else "Unknown error"
                finally:
                    service.config.model_name = original_model

            except Exception as e:
                logger.error(f'Error in generation: {str(e)}', exc_info=True)
                return None, str(e)

        # Run async
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        try:
            result, error = loop.run_until_complete(generate())
        finally:
            loop.close()

        if result:
            logger.info("✅ System prompt generated successfully")
            return jsonify({
                'success': True,
                'system_prompt': result.strip()
            })
        else:
            logger.error(f"❌ Generation failed: {error}")
            return jsonify({
                'success': False,
                'error': f'Failed to generate: {error}',
                'system_prompt': ''
            }), 500

    except Exception as e:
        logger.error(f'❌ Exception in endpoint: {str(e)}', exc_info=True)
        return jsonify({
            'success': False,
            'error': f'Error: {str(e)}',
            'system_prompt': ''
        }), 500

@app.route('/api/generate-questions', methods=['POST'])
def generate_questions():
    """Generate training questions for extracted data"""
    try:
        # Check if any AI service is available
        if not ai_service_manager:
            return jsonify({
                'error': 'Question generation service is not available. Please check AI provider configuration.',
                'success': False
            }), 503

        data = request.json.get('data', [])
        preferred_provider = request.json.get('provider')
        preferred_model = request.json.get('model')

        # Extract AI parameters
        temperature = request.json.get('temperature', 0.7)
        max_tokens = request.json.get('max_tokens', 300)
        top_p = request.json.get('top_p', 0.9)
        system_prompt = request.json.get('system_prompt')

        # Validate required parameters
        if not data:
            return jsonify({
                'error': 'No data provided for question generation',
                'success': False
            }), 400

        # System prompt is optional - will use default if not provided

        if not preferred_model or not preferred_model.strip():
            return jsonify({
                'error': 'Model selection is required. Please select a model before generating questions.',
                'success': False
            }), 400

        # Process all data items
        data_to_process = data
        logger.info(f'📊 Processing all {len(data_to_process)} items')

        logger.info(f'🚀 Starting question generation for {len(data_to_process)} content items')

        # Use AI service manager if available
        if ai_service_manager:
            logger.info('✅ Using AI service manager for question generation')

            # Generate questions using the service manager
            async def generate_questions_async():
                processed_results = []
                successful_count = 0

                for i, item in enumerate(data_to_process):
                    content = item.get('content', '')
                    context = {
                        'section': item.get('section', ''),
                        'subsection': item.get('subsection', ''),
                        'page': item.get('page', '')
                    }

                    # Generate question using AI service manager
                    response = await ai_service_manager.generate_question(content, context, preferred_provider, preferred_model)

                    result_item = item.copy()
                    if response.success:
                        result_item['question'] = response.content
                        result_item['question_generated'] = True
                        result_item['provider_used'] = response.provider.value if response.provider else 'unknown'
                        successful_count += 1
                    else:
                        result_item['question'] = ''
                        result_item['question_generated'] = False
                        result_item['error'] = response.error
                        result_item['provider_used'] = response.provider.value if response.provider else 'unknown'

                    processed_results.append(result_item)

                # Merge processed results back with original data
                final_results = []
                processed_dict = {item['id']: item for item in processed_results}

                for original_item in data:
                    if original_item['id'] in processed_dict:
                        # Use the processed version with new question
                        final_results.append(processed_dict[original_item['id']])
                    else:
                        # Keep the original item unchanged (already had a question)
                        final_results.append(original_item)

                return final_results, successful_count

            # Run async generation
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            try:
                results, successful_count = loop.run_until_complete(generate_questions_async())
            finally:
                loop.close()

            provider_used = 'ai_service_manager'

        else:
            return jsonify({
                'error': 'No AI services available for question generation',
                'success': False
            }), 503

        logger.info(f'🎯 Question generation completed: {successful_count}/{len(data_to_process)} successful')

        # Log the generated questions for debugging
        for i, item in enumerate(results):
            if item.get('question_generated') and not item.get('_has_existing_question'):
                logger.info(f'✅ Generated question {i+1}: {item.get("question", "")[:100]}...')
            elif item.get('_has_existing_question'):
                logger.info(f'📝 Kept existing question {i+1}: {item.get("question", "")[:100]}...')
            else:
                logger.warning(f'❌ Failed to generate question for item {i+1}')

        return jsonify({
            'success': True,
            'data': results,
            'total_items': len(data),
            'successful_generations': successful_count,
            'message': f'Generated {successful_count} new questions (processed {len(data_to_process)} items)',
            'provider_used': provider_used
        })

    except Exception as e:
        logger.error(f'❌ Error generating questions: {str(e)}')
        return jsonify({
            'error': f'Failed to generate questions: {str(e)}',
            'success': False
        }), 500

# Global variable to track question generation progress
question_generation_progress = {
    'is_generating': False,
    'current_item': 0,
    'total_items': 0,
    'progress_percent': 0,
    'current_question': None,
    'generated_questions': [],
    'successful_count': 0,
    'failed_count': 0,
    'retry_count': 0,
    'current_status': 'idle',  # idle, processing, retrying, rate_limited, completed, failed
    'rate_limit_wait': 0,
    'estimated_completion': None
}

@app.route('/api/generate-questions-with-progress', methods=['POST'])
def generate_questions_with_progress():
    """Generate training questions with real-time progress updates"""
    global question_generation_progress

    try:
        # Sync credentials from request headers to credential manager (in-memory only)
        from request_credentials import sync_request_credentials_to_manager
        sync_request_credentials_to_manager(request)

        # Check if any AI service is available
        if not ai_service_manager:
            return jsonify({
                'error': 'Question generation service is not available. Please check AI provider configuration.',
                'success': False
            }), 503

        data = request.json.get('data', [])
        preferred_provider = request.json.get('provider')
        preferred_model = request.json.get('model')
        disable_fallback = request.json.get('disable_fallback', False)  # Default to allow fallback
        generation_mode = request.json.get('generation_mode', 'qa_pair')  # Default to Q&A pair mode

        # Extract AI parameters
        temperature = request.json.get('temperature', 0.7)
        max_tokens = request.json.get('max_tokens', 300)
        top_p = request.json.get('top_p', 0.9)
        system_prompt = request.json.get('system_prompt')

        # Validate required parameters
        if not data:
            return jsonify({
                'error': 'No data provided for question generation',
                'success': False
            }), 400

        # System prompt is optional - will use default if not provided

        if not preferred_model or not preferred_model.strip():
            return jsonify({
                'error': 'Model selection is required. Please select a model before generating questions.',
                'success': False
            }), 400

        # Ensure all data items have unique IDs for tracking
        for i, item in enumerate(data):
            if 'id' not in item:
                item['id'] = f"item_{i}"

        # Process all data items
        data_to_process = data
        logger.info(f'📊 Processing all {len(data_to_process)} items')

        # Initialize progress tracking
        question_generation_progress = {
            'is_generating': True,
            'current_item': 0,
            'total_items': len(data_to_process),
            'progress_percent': 0,
            'current_question': None,
            'generated_questions': [],
            'successful_count': 0,
            'failed_count': 0,
            'retry_count': 0,
            'current_status': 'processing',
            'rate_limit_wait': 0,
            'estimated_completion': None
        }

        logger.info(f'🚀 Starting real-time question generation for {len(data_to_process)} content items')

        # Generate questions one by one with progress updates
        results = []
        successful_count = 0

        # Use AI service manager if available
        if ai_service_manager:
            logger.info('✅ Using AI service manager for real-time generation')

            # Convert preferred_provider string to AIProviderType if specified
            preferred_provider_type = None
            if preferred_provider:
                from ai_service_base import AIProviderType
                for pt in AIProviderType:
                    if pt.value == preferred_provider:
                        preferred_provider_type = pt
                        break

                # Ensure the preferred provider is initialized with request credentials
                if preferred_provider_type:
                    from request_credentials import extract_credentials_from_headers
                    from ai_service_base import AIServiceConfig

                    creds = extract_credentials_from_headers(request, preferred_provider)
                    if creds:
                        # Create config from request credentials
                        config = AIServiceConfig(
                            provider_type=preferred_provider_type,
                            api_key=creds.api_key,
                            endpoint=creds.endpoint,
                            deployment_name=creds.deployment_name,
                            model_name=creds.model_name,
                            api_version=creds.api_version
                        )

                        # Ensure provider is initialized with these credentials
                        logger.info(f"🔧 Ensuring {preferred_provider} is initialized with request credentials")

                        # Run async initialization in event loop
                        loop = asyncio.new_event_loop()
                        asyncio.set_event_loop(loop)
                        try:
                            initialized = loop.run_until_complete(
                                ai_service_manager.ensure_provider_initialized(preferred_provider_type, config)
                            )
                            if initialized:
                                logger.info(f"✅ {preferred_provider} initialized successfully")
                            else:
                                logger.error(f"❌ Failed to initialize {preferred_provider}")
                                if disable_fallback:
                                    return jsonify({
                                        'error': f'Failed to initialize {preferred_provider}. Please check your credentials.',
                                        'success': False
                                    }), 400
                        finally:
                            loop.close()
                    else:
                        logger.warning(f"⚠️ No credentials found in request headers for {preferred_provider}")
                        if disable_fallback:
                            return jsonify({
                                'error': f'No credentials provided for {preferred_provider}. Please configure the provider.',
                                'success': False
                            }), 400

            async def generate_questions_with_progress_async():
                nonlocal successful_count
                processed_results = []

                for i, item in enumerate(data_to_process):
                    # Update progress - show current item being processed
                    question_generation_progress['current_item'] = i + 1
                    question_generation_progress['current_status'] = 'processing'

                    logger.info(f"📝 Processing item {i + 1}/{len(data_to_process)}...")

                    content = item.get('content', '')

                    # DEBUG: Log content details
                    logger.info(f"🔍 Content details for item {i + 1}:")
                    logger.info(f"   Content length: {len(content)}")
                    logger.info(f"   Content preview: {repr(content[:100]) if content else 'EMPTY'}")

                    # Validate content length before processing
                    if not content or len(content.strip()) < 10:
                        logger.warning(f"⚠️ Skipping item {i + 1}: Content too short (length: {len(content.strip()) if content else 0})")
                        question_generation_progress['failed_count'] += 1
                        question_generation_progress['progress_percent'] = int(((i + 1) / len(data_to_process)) * 100)

                        # Add failed result
                        processed_results.append({
                            **item,
                            'question_generated': False,
                            'generation_error': f'Content too short for question generation (minimum 10 characters required, got {len(content.strip()) if content else 0})'
                        })
                        continue

                    context = {
                        'section': item.get('section', ''),
                        'subsection': item.get('subsection', ''),
                        'subsubsection': item.get('subsubsection', ''),
                        'subsubsubsection': item.get('subsubsubsection', '')
                    }

                    # Attempt question-answer generation with enhanced error handling
                    question = None
                    answer = None
                    max_item_retries = 3
                    provider_used = None

                    for retry_attempt in range(max_item_retries):
                        try:
                            question_generation_progress['current_status'] = f'processing_attempt_{retry_attempt + 1}'

                            # Generate question-answer pair using AI service manager
                            response = await ai_service_manager.generate_question_answer_pair(
                                content, context, preferred_provider_type, preferred_model, disable_fallback,
                                temperature, max_tokens, top_p, system_prompt, generation_mode
                            )

                            # DEBUG: Print raw response for debugging
                            print(f"\n🔍 DEBUG - Item {i + 1}, Attempt {retry_attempt + 1}")
                            print(f"📤 Provider: {response.provider.value if response.provider else 'unknown'}")
                            print(f"✅ Success: {response.success}")
                            print(f"📝 Raw Response Content:")
                            print(f"   Type: {type(response.content)}")
                            print(f"   Length: {len(str(response.content)) if response.content else 0}")
                            print(f"   Content: {repr(response.content)}")
                            if response.error:
                                print(f"❌ Error: {response.error}")
                            print("-" * 80)

                            if response.success:
                                # Handle response based on generation mode
                                if generation_mode == 'question_only':
                                    # Question-only mode: use response as question, content as answer
                                    print(f"📝 Question-only mode: Using content as answer")
                                    question = response.content if isinstance(response.content, str) else str(response.content)
                                    answer = content  # Use original content as answer
                                    print(f"✅ Question-only result:")
                                    print(f"   Question: {repr(question)}")
                                    print(f"   Answer: {repr(answer)} (original content)")
                                else:
                                    # Q&A pair mode: parse JSON response
                                    try:
                                        import json
                                        print(f"🔄 Attempting JSON parsing...")

                                        # First try direct parsing if it's already a dict
                                        if isinstance(response.content, dict):
                                            qa_data = response.content
                                            print(f"✅ Response is already a dictionary: {type(qa_data)}")
                                        elif isinstance(response.content, str):
                                            # Try direct JSON parsing first
                                            try:
                                                qa_data = json.loads(response.content)
                                                print(f"✅ Direct JSON parsing successful: {type(qa_data)}")
                                            except json.JSONDecodeError:
                                                # Use improved JSON extraction for malformed responses
                                                print(f"   Direct parsing failed, trying JSON extraction...")
                                                json_str = extract_json_from_text(response.content)
                                                if json_str:
                                                    qa_data = json.loads(json_str)
                                                    print(f"✅ Extracted JSON parsing successful: {type(qa_data)}")
                                                else:
                                                    raise json.JSONDecodeError("No valid JSON found", response.content, 0)
                                        else:
                                            raise TypeError(f"Unexpected response content type: {type(response.content)}")

                                        print(f"📊 Parsed data: {repr(qa_data)}")

                                        if isinstance(qa_data, dict) and 'question' in qa_data and 'answer' in qa_data:
                                            question = qa_data['question']
                                            answer = qa_data['answer']  # Only use parsed answer from JSON
                                            print(f"✅ Valid Q&A pair extracted:")
                                            print(f"   Question: {repr(question)}")
                                            print(f"   Answer: {repr(answer)}")
                                        else:
                                            # If JSON parsing fails, leave answer empty
                                            question = response.content if response.content else None
                                            answer = ''  # Empty answer if JSON parsing fails
                                            print(f"❌ Invalid Q&A structure - missing keys or wrong format")
                                            print(f"   Fallback Question: {repr(question)}")
                                            print(f"   Answer: {repr(answer)} (empty)")
                                    except (json.JSONDecodeError, TypeError) as e:
                                        # If JSON parsing fails, leave answer empty
                                        question = response.content if response.content else None
                                        answer = ''  # Empty answer if JSON parsing fails
                                        print(f"❌ JSON parsing failed: {str(e)}")
                                        print(f"   Fallback Question: {repr(question)}")
                                        print(f"   Answer: {repr(answer)} (empty)")

                                print(f"🎯 Final Result:")
                                print(f"   Question: {repr(question)}")
                                print(f"   Answer: {repr(answer)}")
                                print("=" * 80)

                                provider_used = response.provider.value if response.provider else 'unknown'
                                break  # Success, exit retry loop
                            else:
                                logger.warning(f"⚠️ Item {i + 1}, attempt {retry_attempt + 1}: {response.error}")

                                # Check if this is a provider-specific error with disable_fallback
                                if (disable_fallback and response.provider and
                                    response.metadata and response.metadata.get('disable_fallback')):
                                    # Return immediately with provider-specific error information
                                    question_generation_progress['is_generating'] = False
                                    question_generation_progress['current_status'] = 'failed'

                                    return [], 0, {
                                        'failed_provider': response.provider.value,
                                        'provider_display_name': response.metadata.get('provider_display_name'),
                                        'error_message': response.error
                                    }

                                if retry_attempt < max_item_retries - 1:
                                    question_generation_progress['current_status'] = 'retrying'
                                    question_generation_progress['retry_count'] += 1
                                    await asyncio.sleep(1)  # Brief pause before retry

                        except Exception as e:
                            error_msg = str(e).lower()
                            if '429' in error_msg or 'rate limit' in error_msg:
                                # Rate limit detected
                                question_generation_progress['current_status'] = 'rate_limited'
                                wait_time = min(30, 2 ** retry_attempt)  # Cap at 30 seconds
                                question_generation_progress['rate_limit_wait'] = wait_time

                                logger.warning(f"🚦 Rate limit hit for item {i + 1}, waiting {wait_time}s (attempt {retry_attempt + 1})")
                                await asyncio.sleep(wait_time)
                                question_generation_progress['rate_limit_wait'] = 0
                            else:
                                logger.warning(f"⚠️ Item {i + 1}, attempt {retry_attempt + 1} failed: {e}")
                                if retry_attempt < max_item_retries - 1:
                                    question_generation_progress['current_status'] = 'retrying'
                                    question_generation_progress['retry_count'] += 1
                                    await asyncio.sleep(1)

                    # Create result item
                    result_item = item.copy()
                    result_item['question'] = question
                    result_item['answer'] = answer
                    result_item['question_generated'] = question is not None
                    result_item['provider_used'] = provider_used

                    if question:
                        successful_count += 1
                        question_generation_progress['successful_count'] = successful_count
                        question_generation_progress['current_question'] = question
                        question_generation_progress['generated_questions'].append({
                            'index': i,
                            'question': question,
                            'answer': answer or '',  # Include answer in progress tracking
                            'content': content[:100] + '...',
                            'section': item.get('section', ''),
                            'provider_used': provider_used
                        })
                        logger.info(f"✅ Item {i + 1}: Question-Answer pair generated successfully using {provider_used}")
                    else:
                        question_generation_progress['failed_count'] += 1
                        logger.warning(f"❌ Item {i + 1}: Question generation failed after {max_item_retries} attempts")

                    processed_results.append(result_item)

                    # Update progress percentage after processing item
                    question_generation_progress['progress_percent'] = int(((i + 1) / len(data_to_process)) * 100)
                    logger.info(f"✅ Item {i + 1}/{len(data_to_process)} processed ({question_generation_progress['progress_percent']}%)")

                # Merge processed results back with original data
                final_results = []
                processed_dict = {item['id']: item for item in processed_results}

                for original_item in data:
                    if original_item['id'] in processed_dict:
                        # Use the processed version with new question
                        final_results.append(processed_dict[original_item['id']])
                    else:
                        # Keep the original item unchanged (already had a question)
                        final_results.append(original_item)

                return final_results, successful_count, None

            # Run async generation
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            try:
                results, successful_count, provider_error = loop.run_until_complete(generate_questions_with_progress_async())

                # Check if there was a provider-specific error
                if provider_error:
                    return jsonify({
                        'success': False,
                        'error': provider_error['error_message'],
                        'failed_provider': provider_error['failed_provider'],
                        'provider_display_name': provider_error['provider_display_name']
                    }), 400

            finally:
                loop.close()

            provider_used_display = preferred_provider if preferred_provider else 'ai_service_manager'

        else:
            return jsonify({
                'error': 'No AI services available for question generation',
                'success': False
            }), 503

        # Mark generation as complete
        question_generation_progress['is_generating'] = False
        question_generation_progress['progress_percent'] = 100
        question_generation_progress['current_status'] = 'completed'
        question_generation_progress['rate_limit_wait'] = 0

        logger.info(f'🎯 Real-time question generation completed: {successful_count}/{len(data_to_process)} successful, {question_generation_progress["failed_count"]} failed')

        return jsonify({
            'success': True,
            'data': results,
            'total_items': len(data),
            'successful_generations': successful_count,
            'message': f'Generated {successful_count} new questions (processed {len(data_to_process)} items)',
            'provider_used': provider_used_display
        })

    except Exception as e:
        question_generation_progress['is_generating'] = False
        question_generation_progress['current_status'] = 'failed'
        question_generation_progress['rate_limit_wait'] = 0
        logger.error(f'❌ Error in real-time question generation: {str(e)}')
        return jsonify({
            'error': f'Failed to generate questions: {str(e)}',
            'success': False
        }), 500

@app.route('/api/question-generation-progress', methods=['GET'])
def get_question_generation_progress():
    """Get current question generation progress"""
    return jsonify(question_generation_progress)

@app.route('/api/question-generation-status', methods=['GET'])
def question_generation_status():
    """Get question generation service status"""
    # Wait for initialization to complete (with timeout)
    max_wait_seconds = 10  # Wait up to 10 seconds for initialization
    wait_interval = 0.1  # Check every 100ms
    elapsed = 0

    while not ai_initialization_complete and elapsed < max_wait_seconds:
        time.sleep(wait_interval)
        elapsed += wait_interval

    if not ai_initialization_complete:
        logger.warning(f"AI initialization still in progress after {max_wait_seconds}s timeout")

    # Check if question generation is enabled
    enabled = os.getenv('QUESTION_GENERATION_ENABLED', 'false').lower() == 'true'

    # Check service availability
    service_available = False
    available_providers = []
    primary_provider = None

    if enabled:
        # Check AI service manager
        if ai_service_manager:
            available_providers = [provider.value for provider in ai_service_manager.get_available_providers()]
            service_available = len(available_providers) > 0

            # Get primary provider
            if hasattr(ai_service_manager, '_primary_provider') and ai_service_manager._primary_provider:
                primary_provider = ai_service_manager._primary_provider.value

    return jsonify({
        'enabled': enabled,
        'service_available': service_available,
        'configuration_valid': service_available,
        'available_providers': available_providers,
        'primary_provider': primary_provider,
        'total_providers': len(available_providers),
        'ai_service_manager_available': ai_service_manager is not None,
        'initialization_complete': ai_initialization_complete
    })

@app.route('/api/ai-providers/quick-status', methods=['GET'])
def get_ai_providers_quick_status():
    """Get quick AI provider status without waiting for initialization - optimized for polling"""
    try:
        from request_credentials import extract_all_credentials_from_headers

        providers = {}

        # Provider display names
        provider_names = {
            'openai': 'OpenAI',
            'google_gemini': 'Google Gemini',
            'lm_studio': 'LM Studio',
            'ollama': 'Ollama'
        }

        # Get credentials from request headers (client-side storage)
        request_credentials = extract_all_credentials_from_headers(request)
        all_provider_ids = ['openai', 'google_gemini', 'lm_studio', 'ollama']

        # Check AI service manager for initialized providers (no waiting)
        initialized_providers = {}
        primary_provider = None
        if ai_service_manager and ai_initialization_complete:
            provider_status = ai_service_manager.get_provider_status()
            primary_provider = getattr(ai_service_manager, '_primary_provider', None)

            for provider_id, status in provider_status.items():
                initialized_providers[provider_id] = status

        # Build provider list
        for provider_id in all_provider_ids:
            credentials = request_credentials.get(provider_id)
            has_credentials = credentials is not None and credentials.api_key is not None

            # Check if provider requires API key
            requires_api_key = provider_id in ['openai', 'google_gemini']

            if provider_id in initialized_providers:
                # Provider is initialized - use its status
                status = initialized_providers[provider_id]
                providers[provider_id] = {
                    'name': provider_names.get(provider_id, provider_id.replace('_', ' ').title()),
                    'available': status['status'] == 'available',
                    'isDefault': primary_provider and primary_provider.value == provider_id,
                    'status': status['status'],
                    'models': [model['name'] for model in status['available_models']] if status['available_models'] else [],
                    'last_error': status['last_error'],
                    'has_credentials': has_credentials
                }
            elif has_credentials:
                # Provider has credentials but not initialized yet
                providers[provider_id] = {
                    'name': provider_names.get(provider_id, provider_id.replace('_', ' ').title()),
                    'available': False,
                    'isDefault': False,
                    'status': 'configured',
                    'models': [],
                    'last_error': None,
                    'has_credentials': True
                }
            else:
                # Provider has no credentials
                providers[provider_id] = {
                    'name': provider_names.get(provider_id, provider_id.replace('_', ' ').title()),
                    'available': False,
                    'isDefault': False,
                    'status': 'not_configured',
                    'models': [],
                    'last_error': None,
                    'has_credentials': False
                }

        return jsonify({
            'success': True,
            'data': providers,
            'initialization_complete': ai_initialization_complete
        })

    except Exception as e:
        logger.error(f'Error getting AI providers quick status: {str(e)}')
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/ai-providers', methods=['GET'])
def get_ai_providers():
    """Get available AI providers with status based on credentials and initialization"""
    try:
        from request_credentials import extract_all_credentials_from_headers

        # Wait for initialization to complete (with timeout)
        max_wait_seconds = 10  # Wait up to 10 seconds for initialization
        wait_interval = 0.1  # Check every 100ms
        elapsed = 0

        while not ai_initialization_complete and elapsed < max_wait_seconds:
            time.sleep(wait_interval)
            elapsed += wait_interval

        if not ai_initialization_complete:
            logger.warning(f"AI initialization still in progress after {max_wait_seconds}s timeout")

        providers = {}

        # Provider display names
        provider_names = {
            'openai': 'OpenAI',
            'google_gemini': 'Google Gemini',
            'lm_studio': 'LM Studio',
            'ollama': 'Ollama'
        }

        # Get credentials from request headers (client-side storage)
        request_credentials = extract_all_credentials_from_headers(request)
        all_provider_ids = ['openai', 'google_gemini', 'lm_studio', 'ollama']

        # Check AI service manager for initialized providers
        initialized_providers = {}
        primary_provider = None
        if ai_service_manager:
            provider_status = ai_service_manager.get_provider_status()
            primary_provider = getattr(ai_service_manager, '_primary_provider', None)

            for provider_id, status in provider_status.items():
                initialized_providers[provider_id] = status

        # Build provider list including both initialized and credential-only providers
        for provider_id in all_provider_ids:
            credentials = request_credentials.get(provider_id)
            has_credentials = credentials is not None and credentials.api_key is not None

            # Check if provider requires API key
            requires_api_key = provider_id in ['openai', 'google_gemini']
            has_valid_credentials = False

            if requires_api_key:
                has_valid_credentials = has_credentials
            else:
                # Local providers (LM Studio, Ollama) don't require API key
                has_valid_credentials = True

            if provider_id in initialized_providers:
                # Provider is initialized - use its status
                status = initialized_providers[provider_id]
                providers[provider_id] = {
                    'name': provider_names.get(provider_id, provider_id.replace('_', ' ').title()),
                    'available': status['status'] == 'available',
                    'isDefault': primary_provider and primary_provider.value == provider_id,
                    'status': status['status'],
                    'models': [model['name'] for model in status['available_models']] if status['available_models'] else [],
                    'last_error': status['last_error'],
                    'has_credentials': has_credentials
                }
            elif has_credentials:
                # Provider has credentials but not initialized yet
                providers[provider_id] = {
                    'name': provider_names.get(provider_id, provider_id.replace('_', ' ').title()),
                    'available': False,
                    'isDefault': False,
                    'status': 'configured',  # Has credentials but not tested
                    'models': [],
                    'last_error': None,
                    'has_credentials': True
                }
            else:
                # Provider has no credentials
                providers[provider_id] = {
                    'name': provider_names.get(provider_id, provider_id.replace('_', ' ').title()),
                    'available': False,
                    'isDefault': False,
                    'status': 'not_configured',
                    'models': [],
                    'last_error': None,
                    'has_credentials': False
                }

        return jsonify({
            'success': True,
            'data': providers,
            'initialization_complete': ai_initialization_complete
        })

    except Exception as e:
        logger.error(f'Error getting AI providers: {str(e)}')
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/ai-providers/<provider_id>/test', methods=['POST'])
def test_ai_provider(provider_id):
    """
    Test specific AI provider connection with on-demand initialization.

    This endpoint:
    1. Receives credentials from request headers (client-side storage)
    2. Creates a temporary service instance
    3. Tests the connection with a "hello" message
    4. Returns detailed status information
    """
    try:
        from request_credentials import get_provider_credential
        from ai_service_base import AIProviderType
        from ai_config_loader import create_config_from_request_credentials

        # Get credentials from request headers (client-side storage)
        credentials = get_provider_credential(request, provider_id)

        # Local providers (Ollama, LM Studio) don't require API keys
        local_providers = ['ollama', 'lm_studio']
        is_local = provider_id in local_providers

        if not credentials:
            error_msg = f'No endpoint configured for {provider_id}' if is_local else f'No credentials provided for {provider_id}. Please configure your API key.'
            return jsonify({
                'success': False,
                'provider': provider_id,
                'available': False,
                'error': error_msg
            })

        # Convert provider_id to AIProviderType
        provider_type = None
        for pt in AIProviderType:
            if pt.value == provider_id:
                provider_type = pt
                break

        if not provider_type:
            return jsonify({
                'success': False,
                'provider': provider_id,
                'available': False,
                'error': f'Unsupported provider: {provider_id}'
            }), 400

        # Create a temporary service instance for testing
        async def test_provider_standalone():
            """Test provider connection without requiring full initialization"""
            try:
                # Create config from request credentials
                config = create_config_from_request_credentials(provider_id, credentials)
                if not config:
                    return {
                        'success': False,
                        'error': 'Failed to create configuration from provided credentials'
                    }

                # Create temporary service instance
                if ai_service_manager:
                    service = ai_service_manager._create_service(provider_type, config)

                    # Initialize the service
                    logger.info(f"Testing {provider_id} connection...")
                    init_success = await service.initialize()
                    if not init_success:
                        error_msg = service.last_error or 'Initialization failed'
                        logger.warning(f"{provider_id} initialization failed: {error_msg}")
                        return {
                            'success': False,
                            'error': error_msg
                        }

                    # Test connection with a simple message
                    logger.info(f"Sending test message to {provider_id}...")
                    test_result = await service.test_connection()

                    if test_result.success:
                        logger.info(f"✅ {provider_id} connection test successful")
                        return {
                            'success': True,
                            'message': f'{provider_id} is working correctly',
                            'response_time': test_result.response_time if hasattr(test_result, 'response_time') else None
                        }
                    else:
                        logger.warning(f"❌ {provider_id} connection test failed: {test_result.error}")
                        return {
                            'success': False,
                            'error': test_result.error or 'Connection test failed'
                        }
                else:
                    return {
                        'success': False,
                        'error': 'AI service manager not available'
                    }

            except Exception as e:
                logger.error(f"Error testing {provider_id}: {e}", exc_info=True)
                return {
                    'success': False,
                    'error': str(e)
                }

        # Run async test with timeout
        try:
            import concurrent.futures
            timeout = 60 if is_local else 30  # Longer timeout for local providers

            with concurrent.futures.ThreadPoolExecutor() as executor:
                future = executor.submit(asyncio.run, test_provider_standalone())
                result = future.result(timeout=timeout)
        except concurrent.futures.TimeoutError:
            logger.error(f"Timeout testing {provider_id}")
            result = {
                'success': False,
                'error': f'Connection test timed out after {timeout} seconds'
            }
        except Exception as e:
            logger.error(f"Error running async test for {provider_id}: {e}")
            result = {
                'success': False,
                'error': str(e)
            }

        # Return detailed response
        response = {
            'success': result['success'],
            'provider': provider_id,
            'available': result['success'],
            'error': result.get('error'),
            'message': result.get('message') or ('Connection successful!' if result['success'] else None)
        }

        if result.get('response_time'):
            response['response_time'] = result['response_time']

        return jsonify(response)

    except Exception as e:
        logger.error(f'Error testing AI provider {provider_id}: {str(e)}', exc_info=True)
        return jsonify({
            'success': False,
            'provider': provider_id,
            'available': False,
            'error': str(e)
        }), 500

@app.route('/api/ai-providers/<provider_id>/models', methods=['GET'])
def get_provider_models(provider_id):
    """Get available models for a specific AI provider

    For Ollama and LM Studio: Returns dynamically discovered models
    For OpenAI, Anthropic, and other cloud providers: Returns empty list (use text input)
    """
    try:
        from request_credentials import get_provider_credential
        from ai_service_base import AIProviderType
        from ai_config_loader import create_config_from_request_credentials

        # Debug: Log all headers
        logger.info(f"🔍 Getting models for provider: {provider_id}")
        logger.info(f"📋 Request headers: {dict(request.headers)}")

        # Wait for initialization to complete (with timeout)
        max_wait_seconds = 10
        wait_interval = 0.1
        elapsed = 0

        while not ai_initialization_complete and elapsed < max_wait_seconds:
            time.sleep(wait_interval)
            elapsed += wait_interval

        # Providers that support dynamic model discovery
        dynamic_discovery_providers = ['ollama', 'lm_studio']

        if ai_service_manager:
            # Get provider status which includes available models
            provider_status = ai_service_manager.get_provider_status()

            if provider_id in provider_status:
                models = provider_status[provider_id].get('available_models', [])

                # For providers without dynamic discovery, return empty list
                # Frontend will show text input instead of dropdown
                if provider_id not in dynamic_discovery_providers:
                    return jsonify({
                        'success': True,
                        'data': [],
                        'provider': provider_id,
                        'use_manual_input': True,
                        'message': 'This provider requires manual model name input'
                    })

                return jsonify({
                    'success': True,
                    'data': models,
                    'provider': provider_id,
                    'use_manual_input': False
                })
            else:
                # Provider not initialized yet - try to get credentials from headers
                # and create a temporary service to discover models
                if provider_id in dynamic_discovery_providers:
                    credentials = get_provider_credential(request, provider_id)
                    logger.info(f"🔑 Credentials extracted for {provider_id}: {credentials}")

                    if credentials:
                        # Try to discover models with provided credentials
                        async def discover_models_standalone():
                            try:
                                # Convert provider_id to AIProviderType
                                provider_type = None
                                for pt in AIProviderType:
                                    if pt.value == provider_id:
                                        provider_type = pt
                                        break

                                if not provider_type:
                                    return {'success': False, 'models': []}

                                # Create config from request credentials
                                config = create_config_from_request_credentials(provider_id, credentials)
                                if not config:
                                    return {'success': False, 'models': []}

                                # Create temporary service instance
                                service = ai_service_manager._create_service(provider_type, config)

                                # For LM Studio and Ollama, we only need to discover models
                                # We don't need to test the chat endpoint for model discovery
                                if provider_id in ['lm_studio', 'ollama']:
                                    # Create session for model discovery only
                                    connector = aiohttp.TCPConnector(
                                        limit=10,
                                        limit_per_host=5,
                                        enable_cleanup_closed=True
                                    )
                                    service._session = aiohttp.ClientSession(
                                        connector=connector,
                                        timeout=aiohttp.ClientTimeout(total=service.config.timeout),
                                        headers={"Content-Type": "application/json"}
                                    )

                                    # Discover models without full initialization
                                    await service._discover_models()

                                    # Get available models
                                    available_models = await service.get_available_models()
                                    logger.info(f"📋 Discovered {len(available_models)} models from {provider_id}")

                                    # Cleanup session
                                    await service.cleanup()

                                    models = [
                                        {
                                            'name': model.name,
                                            'display_name': model.display_name,
                                            'description': model.description,
                                            'max_tokens': model.max_tokens
                                        }
                                        for model in available_models
                                    ]

                                    return {'success': True, 'models': models}
                                else:
                                    # For cloud providers, use full initialization
                                    init_success = await service.initialize()
                                    if not init_success:
                                        logger.error(f"Failed to initialize {provider_id} service")
                                        return {'success': False, 'models': []}

                                    # Get available models using the async method
                                    available_models = await service.get_available_models()
                                    logger.info(f"📋 Got {len(available_models)} models from {provider_id} service")

                                    models = [
                                        {
                                            'name': model.name,
                                            'display_name': model.display_name,
                                            'description': model.description,
                                            'max_tokens': model.max_tokens
                                        }
                                        for model in available_models
                                    ]

                                    # Cleanup
                                    await service.cleanup()

                                    return {'success': True, 'models': models}
                            except Exception as e:
                                logger.error(f"Error discovering models for {provider_id}: {e}")
                                return {'success': False, 'models': []}

                        # Run async discovery
                        try:
                            import concurrent.futures
                            with concurrent.futures.ThreadPoolExecutor() as executor:
                                future = executor.submit(asyncio.run, discover_models_standalone())
                                result = future.result(timeout=30)

                                if result['success']:
                                    return jsonify({
                                        'success': True,
                                        'data': result['models'],
                                        'provider': provider_id,
                                        'use_manual_input': False
                                    })
                        except Exception as e:
                            logger.error(f"Error running async model discovery for {provider_id}: {e}")

                    # No credentials or discovery failed - return empty list
                    return jsonify({
                        'success': True,
                        'data': [],
                        'provider': provider_id,
                        'use_manual_input': False,
                        'message': f'{provider_id} not configured yet. Please configure the endpoint first.'
                    })
                else:
                    return jsonify({
                        'success': False,
                        'error': f'Provider {provider_id} not found'
                    }), 404

        # For providers that don't support dynamic discovery, return empty list
        # This includes: openai, anthropic, google_gemini, xai_grok, deepseek, azure_openai
        if provider_id not in dynamic_discovery_providers:
            return jsonify({
                'success': True,
                'data': [],
                'provider': provider_id,
                'use_manual_input': True,
                'message': 'This provider requires manual model name input'
            })

        # For Ollama and LM Studio without service manager, return fallback models
        # These will be replaced by actual discovered models when service is initialized
        fallback_models = {
            'lm_studio': [
                {
                    'name': 'local-model',
                    'display_name': 'Local Model',
                    'description': 'Local model running in LM Studio',
                    'max_tokens': 4096
                }
            ],
            'ollama': [
                {
                    'name': 'llama2',
                    'display_name': 'Llama 2',
                    'description': 'Default Llama 2 model',
                    'max_tokens': 4096
                }
            ]
        }

        if provider_id in fallback_models:
            return jsonify({
                'success': True,
                'data': fallback_models[provider_id],
                'provider': provider_id,
                'use_manual_input': False,
                'message': 'Fallback models - actual models will be discovered when service is running'
            })

        return jsonify({
            'success': False,
            'error': f'Provider {provider_id} not available'
        }), 404

    except Exception as e:
        logger.error(f'Error getting models for provider {provider_id}: {str(e)}')
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/credentials', methods=['GET'])
def get_all_credentials():
    """
    DEPRECATED: Backend credential storage is disabled.
    All credentials are stored client-side in browser localStorage.
    """
    return jsonify({
        'success': False,
        'error': 'Backend credential storage is disabled. Credentials are stored client-side only (browser localStorage).',
        'message': '🔒 For security and user isolation, all API keys are stored in your browser only.'
    }), 410  # 410 Gone - endpoint is permanently disabled


@app.route('/api/credentials/<provider_id>', methods=['GET'])
def get_provider_credentials(provider_id):
    """
    DEPRECATED: Backend credential storage is disabled.
    All credentials are stored client-side in browser localStorage.
    """
    return jsonify({
        'success': False,
        'error': 'Backend credential storage is disabled. Credentials are stored client-side only (browser localStorage).',
        'message': '🔒 For security and user isolation, all API keys are stored in your browser only.'
    }), 410  # 410 Gone - endpoint is permanently disabled


@app.route('/api/credentials/<provider_id>', methods=['POST', 'PUT'])
def save_provider_credentials(provider_id):
    """
    DEPRECATED: Backend credential storage is disabled.
    All credentials are stored client-side in browser localStorage.
    """
    return jsonify({
        'success': False,
        'error': 'Backend credential storage is disabled. Credentials are stored client-side only (browser localStorage).',
        'message': '🔒 For security and user isolation, all API keys are stored in your browser only. Use the frontend API Key Configuration dialog.'
    }), 410  # 410 Gone - endpoint is permanently disabled


@app.route('/api/credentials/<provider_id>', methods=['DELETE'])
def delete_provider_credentials(provider_id):
    """
    DEPRECATED: Backend credential storage is disabled.
    All credentials are stored client-side in browser localStorage.
    """
    return jsonify({
        'success': False,
        'error': 'Backend credential storage is disabled. Credentials are stored client-side only (browser localStorage).',
        'message': '🔒 For security and user isolation, all API keys are stored in your browser only. Use the frontend API Key Configuration dialog.'
    }), 410  # 410 Gone - endpoint is permanently disabled


@app.route('/api/test', methods=['GET'])
def test_endpoint():
    """Test endpoint for debugging"""
    return jsonify({
        'message': 'Backend is working!',
        'upload_folder': app.config['UPLOAD_FOLDER'],
        'max_file_size': app.config['MAX_CONTENT_LENGTH'],
        'question_generation_enabled': ai_service_manager is not None
    })


















def cleanup_ai_services():
    """Clean up AI services on shutdown."""
    global ai_service_manager
    if ai_service_manager:
        try:
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            loop.run_until_complete(ai_service_manager.cleanup())
            loop.close()
            logger.info("AI services cleaned up successfully")
        except Exception as e:
            logger.error(f"Error cleaning up AI services: {e}")

# Register cleanup handler
import atexit
atexit.register(cleanup_ai_services)

if __name__ == '__main__':
    print("🚀 Starting LLM Data Generator Backend...")
    print("📡 API will be available at: http://localhost:5000")
    print("🔗 Health check: http://localhost:5000/api/health")
    print("📄 PDF processing: POST http://localhost:5000/api/process-pdf")
    print("📊 Excel download: POST http://localhost:5000/api/download-excel")
    print("🤖 Question generation: POST http://localhost:5000/api/generate-questions")
    print("📊 Question gen status: GET http://localhost:5000/api/question-generation-status")
    print("📤 Dataset export: POST http://localhost:5000/api/export-dataset")
    print("🧪 Test endpoint: GET http://localhost:5000/api/test")

    try:
        app.run(debug=True, host='0.0.0.0', port=5000)
    finally:
        cleanup_ai_services()
